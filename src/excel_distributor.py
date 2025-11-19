# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import sys
from datetime import datetime

# 출력 인코딩 설정 (EXE 환경에서는 콘솔이 없을 수 있으므로 안전하게 처리)
if sys.platform == 'win32' and hasattr(sys.stdout, 'buffer') and sys.stdout.buffer is not None:
    try:
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
    except (AttributeError, ValueError):
        # EXE 환경이나 콘솔이 없는 경우 무시
        pass

def load_distribution_rules(file_path):
    """분배 규칙 파일을 읽어서 딕셔너리로 변환"""
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        rules = {}
        
        # 컬럼명 확인 및 정규화
        columns = df.columns.tolist()
        print(f"분배 규칙 파일 컬럼: {columns}")
        
        # 분배 규칙 매핑 생성
        current_product_code = None
        for idx, row in df.iterrows():
            # 다양한 컬럼명 시도
            product_code_col = None
            for col in columns:
                if '출고' in str(col) or '합포장' in str(col) or 'BOX' in str(col).upper():
                    product_code_col = col
                    break
            
            if product_code_col:
                product_code = row.get(product_code_col, '')
            else:
                product_code = row.iloc[0] if len(columns) > 0 else ''
            
            if pd.notna(product_code) and str(product_code).strip().lower() not in ['nan', 'none', '']:
                current_product_code = str(product_code).strip()
                if current_product_code not in rules:
                    rules[current_product_code] = []
            
            if current_product_code:
                # 구성 상품코드 찾기
                component_code_col = None
                for col in columns:
                    if '구성' in str(col) and '상품' in str(col) and '코드' in str(col):
                        component_code_col = col
                        break
                
                # 탑셀러 코드 찾기
                seller_code_col = None
                for col in columns:
                    if '탑셀러' in str(col) or 'seller' in str(col).lower():
                        seller_code_col = col
                        break
                
                # EA 수량 찾기
                ea_col = None
                for col in columns:
                    if 'EA' in str(col) and '수량' in str(col):
                        ea_col = col
                        break
                
                component_code = row.get(component_code_col, '') if component_code_col else ''
                component_seller_code = row.get(seller_code_col, '') if seller_code_col else ''
                ea_quantity = row.get(ea_col, '') if ea_col else ''
                
                if pd.notna(component_code) and str(component_code).strip().lower() not in ['nan', 'none', '']:
                    # EA 수량 파싱
                    ea_value = 0
                    if pd.notna(ea_quantity):
                        ea_str = str(ea_quantity).replace(' EA', '').replace('EA', '').strip()
                        try:
                            ea_value = int(ea_str)
                        except:
                            try:
                                ea_value = int(float(ea_str))
                            except:
                                ea_value = 0
                    
                    rules[current_product_code].append({
                        'component_code': str(component_code).strip(),
                        'seller_code': str(component_seller_code).strip() if pd.notna(component_seller_code) else '',
                        'ea_quantity': ea_value
                    })
        
        return rules
    except Exception as e:
        print(f"분배 규칙 파일 읽기 오류: {e}")
        import traceback
        traceback.print_exc()
        return {}

def match_product_to_rule(product_name, rules):
    """상품명을 분배 규칙과 매칭"""
    if pd.isna(product_name):
        return None, None
    
    # 상품명에서 키워드 추출
    product_name_upper = str(product_name).upper()
    
    # 규칙 키워드와 매칭 (우선순위 순서대로 체크)
    # 1. Profutura Pre + IF 매칭 (가장 구체적)
    if 'PROFUTURA' in product_name_upper and 'PRE' in product_name_upper:
        for rule_key, components in rules.items():
            rule_key_upper = str(rule_key).upper()
            if '3PF PRE' in rule_key_upper or '3PFPRE' in rule_key_upper or 'PRE' in rule_key_upper:
                return rule_key, components
    
    # 2. Pronutra 1 + IF 매칭
    if 'PRONUTRA' in product_name_upper and ('1' in product_name_upper or 'VP 1' in product_name_upper):
        for rule_key, components in rules.items():
            rule_key_upper = str(rule_key).upper()
            if '3PN1+IF' in rule_key_upper or '3PN 1+IF' in rule_key_upper:
                return rule_key, components
    
    # 3. Pronutra 2 + FO 매칭
    if 'PRONUTRA' in product_name_upper and ('2' in product_name_upper or 'VP 2' in product_name_upper):
        for rule_key, components in rules.items():
            rule_key_upper = str(rule_key).upper()
            if '3PN2+FO' in rule_key_upper or '3PN 2+FO' in rule_key_upper:
                return rule_key, components
    
    # 4. Profutura 1 + IF 매칭
    if 'PROFUTURA' in product_name_upper and '1' in product_name_upper and 'PRE' not in product_name_upper:
        for rule_key, components in rules.items():
            rule_key_upper = str(rule_key).upper()
            if ('3PF 1+IF' in rule_key_upper or '3PF1+IF' in rule_key_upper) and 'PRE' not in rule_key_upper:
                return rule_key, components
    
    # 5. Profutura 2 + FO 매칭
    if 'PROFUTURA' in product_name_upper and '2' in product_name_upper and 'PRE' not in product_name_upper:
        for rule_key, components in rules.items():
            rule_key_upper = str(rule_key).upper()
            if ('3PF 2+FO' in rule_key_upper or '3PF2+FO' in rule_key_upper) and 'PRE' not in rule_key_upper:
                return rule_key, components
    
    return None, None

def distribute_order(order_row, components, quantity_multiplier):
    """주문을 구성 상품으로 분배"""
    distributed_rows = []
    
    for component in components:
        ea_quantity = component['ea_quantity']
        total_ea = ea_quantity * quantity_multiplier
        
        # 새로운 행 생성 (딕셔너리로 변환)
        if isinstance(order_row, pd.Series):
            new_row = order_row.to_dict()
        else:
            new_row = dict(order_row)
        
        new_row['구성 상품코드'] = component['component_code']
        new_row['구성 상품 탑셀러 코드'] = component['seller_code']
        new_row['EA 수량'] = total_ea
        
        # 원본 상품명 찾기
        product_name_col = None
        for col in order_row.index if isinstance(order_row, pd.Series) else order_row.keys():
            if '상품명' in str(col) or 'Commodity' in str(col):
                product_name_col = col
                break
        
        if product_name_col:
            new_row['원본 상품명'] = order_row.get(product_name_col, '')
        
        # 원본 구매 수량 찾기
        quantity_col = None
        for col in order_row.index if isinstance(order_row, pd.Series) else order_row.keys():
            if '구매 수량' in str(col) or 'Quantity' in str(col):
                quantity_col = col
                break
        
        if quantity_col:
            new_row['원본 구매 수량'] = order_row.get(quantity_col, quantity_multiplier)
        else:
            new_row['원본 구매 수량'] = quantity_multiplier
        
        distributed_rows.append(new_row)
    
    return distributed_rows

def process_orders(input_file, rules_file, output_file):
    """주문 파일을 읽어서 분배 규칙에 따라 분배"""
    try:
        # 원본 주문 데이터 읽기
        print("원본 주문 데이터 읽는 중...")
        df_orders = pd.read_excel(input_file, engine='openpyxl')
        print(f"원본 데이터 컬럼: {list(df_orders.columns)}")
        print(f"원본 데이터 행 수: {len(df_orders)}")
        
        # 분배 규칙 읽기
        print("\n분배 규칙 읽는 중...")
        rules = load_distribution_rules(rules_file)
        
        if not rules:
            print("경고: 분배 규칙을 찾을 수 없습니다.")
            return
        
        print(f"\n분배 규칙 {len(rules)}개 로드됨:")
        for key, components in rules.items():
            print(f"  - {key}: {len(components)}개 구성품")
        
        # 분배된 결과 저장할 리스트
        distributed_rows = []
        
        # 상품명 컬럼 찾기
        product_name_col = None
        for col in df_orders.columns:
            if '상품명' in str(col) or 'Commodity' in str(col):
                product_name_col = col
                break
        
        # 구매 수량 컬럼 찾기
        quantity_col = None
        for col in df_orders.columns:
            if '구매 수량' in str(col) or 'Quantity' in str(col):
                quantity_col = col
                break
        
        if not product_name_col:
            print("오류: 상품명 컬럼을 찾을 수 없습니다.")
            return
        
        # 각 주문 처리
        print("\n주문 분배 처리 중...")
        for idx, row in df_orders.iterrows():
            product_name = row.get(product_name_col, '')
            purchase_quantity = row.get(quantity_col, 1) if quantity_col else 1
            
            # 수량 정수 변환
            try:
                purchase_quantity = int(float(purchase_quantity)) if pd.notna(purchase_quantity) else 1
            except:
                purchase_quantity = 1
            
            # 분배 규칙 찾기
            rule_key, components = match_product_to_rule(product_name, rules)
            
            if rule_key and components:
                # 주문 분배
                distributed = distribute_order(row, components, purchase_quantity)
                distributed_rows.extend(distributed)
                print(f"주문 {idx+1}: {product_name} -> {rule_key} 분배 완료 ({len(distributed)}개 행 생성)")
            else:
                # 매칭되는 규칙이 없으면 원본 그대로 추가
                new_row = row.to_dict()
                new_row['구성 상품코드'] = ''
                new_row['구성 상품 탑셀러 코드'] = ''
                new_row['EA 수량'] = purchase_quantity
                new_row['원본 상품명'] = product_name
                new_row['원본 구매 수량'] = purchase_quantity
                distributed_rows.append(new_row)
                print(f"주문 {idx+1}: {product_name} -> 규칙 없음 (원본 유지)")
        
        # 결과 데이터프레임 생성
        df_result = pd.DataFrame(distributed_rows)
        
        # 결과를 엑셀 파일로 저장
        print(f"\n결과 파일 저장 중: {output_file}")
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_result.to_excel(writer, index=False, sheet_name='분배 결과')
        
        # 음영표시 적용
        print("음영표시 적용 중...")
        apply_cell_colors(output_file)
        
        print(f"\n완료! 결과 파일: {output_file}")
        print(f"총 {len(distributed_rows)}개 행 생성됨")
        
    except Exception as e:
        print(f"오류 발생: {e}")
        import traceback
        traceback.print_exc()

def apply_cell_colors(file_path):
    """엑셀 파일에 음영표시 적용 - 분배된 행만 음영처리"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 음영 색상 정의 (연한 회색)
        fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # 헤더 행 음영표시
        for cell in ws[1]:
            cell.fill = fill
        
        # "구성 상품코드" 컬럼 찾기
        header_row = ws[1]
        component_code_col_idx = None
        
        for idx, cell in enumerate(header_row, 1):
            cell_value = str(cell.value) if cell.value else ''
            if '구성 상품코드' in cell_value:
                component_code_col_idx = idx
                break
        
        # 분배된 행만 음영처리 (구성 상품코드가 있는 행)
        if component_code_col_idx:
            for row_idx in range(2, ws.max_row + 1):
                component_cell = ws.cell(row=row_idx, column=component_code_col_idx)
                component_value = str(component_cell.value).strip() if component_cell.value else ''
                
                # 구성 상품코드가 있으면 (비어있지 않으면) 해당 행 전체 음영처리
                if component_value and component_value.lower() not in ['nan', 'none', '']:
                    # 해당 행의 모든 셀에 음영 적용
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = fill
        
        wb.save(file_path)
        print("음영표시 적용 완료")
        
    except Exception as e:
        print(f"음영표시 적용 오류: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # 파일 경로 설정
    input_file = "Krone Order_20251106.xlsx"
    rules_file = "손으로나누는작업.xlsx"
    
    # 출력 파일명 생성 (타임스탬프 포함)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"분배결과_{timestamp}.xlsx"
    
    # 파일 존재 확인
    if not os.path.exists(input_file):
        print(f"오류: 원본 파일을 찾을 수 없습니다: {input_file}")
    elif not os.path.exists(rules_file):
        print(f"오류: 분배 규칙 파일을 찾을 수 없습니다: {rules_file}")
    else:
        process_orders(input_file, rules_file, output_file)

